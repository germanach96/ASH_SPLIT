"""Limpieza de 'Ashford split 2.xlsx' -> relaciones a nivel de codigo.

Uso:  python3 clean_bom_rate.py ["Ashford split 2.xlsx"]

El fichero es una foto de master data, no un historico. Se colapsa a:

  BOM  -> aristas padre/hijo unicas a nivel de codigo de material.
          Se descartan version, posicion y fecha de validez.
  RATE -> pares producto/maquina unicos, conservando la distincion de maquina
          y la etapa productiva. Se descarta la version de ruta.

El fichero se reescribe en sitio; el original queda en el historial de git.
"""

import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FICHERO = sys.argv[1] if len(sys.argv) > 1 else "Ashford split 2.xlsx"

# Productos con ruta pero sin lista de materiales propia: se retiran de RATE.
# 99040009087 no aparece en ninguna BOM; 99200033777 solo figura como componente.
PRODUCTOS_SIN_BOM = ["99040009087", "99200033777"]


def limpiar_bom(bom):
    """BOM -> pares (ParentID, ComponentID) unicos."""
    antes = len(bom)

    # 1. Registros '/out': ids de 5 segmentos donde el numero de operacion cayo
    #    en ComponentID. Son salidas de operacion, no lineas de consumo.
    es_out = bom.BOMElementId.str.count("/") + 1 == 5
    bom = bom[~es_out]
    print(f"  - {es_out.sum()} filas '/out' eliminadas "
          f"(materiales fantasma '0010' y '0040')")

    # 2. Colapso a relacion de codigos. Esto absorbe de golpe los duplicados por
    #    version, por posicion y por fecha de validez: las 8 lineas repetidas con
    #    distinta fecha, las 316 colisiones de posicion y las 8.412 repeticiones
    #    del mismo componente en varias posiciones.
    bom = (bom[["ParentID", "ComponentID"]]
           .drop_duplicates()
           .sort_values(["ParentID", "ComponentID"], kind="stable")
           .reset_index(drop=True))
    print(f"  - {antes} filas -> {len(bom)} aristas padre/hijo unicas")
    return bom


def limpiar_rate(rate):
    """RATE -> pares (ProductID, MachineId) unicos, con la etapa productiva."""
    antes = len(rate)

    # 1. Productos sin BOM propia.
    fuera = rate.ProductID.isin(PRODUCTOS_SIN_BOM)
    rate = rate[~fuera]
    print(f"  - {fuera.sum()} filas eliminadas: {PRODUCTOS_SIN_BOM}")

    # 2. La operacion va dentro del OperationId compuesto y esta determinada al
    #    100% por la maquina (P05M -> 0040, P05P -> 0010). Se extrae a su propia
    #    columna y se descarta la version de ruta.
    rate = rate.assign(OperationNo=rate.OperationId.str.split("/").str[3])

    conflictos = (rate.groupby(["MachineId", "ProductID"]).OperationNo.nunique() > 1).sum()
    assert conflictos == 0, f"{conflictos} pares maquina/producto con operacion ambigua"

    rate = (rate[["MachineId", "ProductID", "OperationNo"]]
            .drop_duplicates()
            .sort_values(["ProductID", "MachineId"], kind="stable")
            .reset_index(drop=True))
    print(f"  - {antes} filas -> {len(rate)} pares producto/maquina unicos")
    return rate


def dar_formato(fichero):
    """Arial en todo el libro, cabecera destacada y columnas legibles."""
    wb = load_workbook(fichero)

    normal = wb._named_styles["Normal"]
    normal.font = Font(name="Arial", size=10)

    cabecera = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="44546A")

    for ws in wb.worksheets:
        for celda in ws[1]:
            celda.font = cabecera
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center")
        for i, col in enumerate(ws.iter_cols(max_row=1), start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(str(col[0].value)) + 6, 16)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(fichero)


def main():
    bom = pd.read_excel(FICHERO, sheet_name="BOM", dtype=str)
    rate = pd.read_excel(FICHERO, sheet_name="RATE", dtype=str)

    print("BOM")
    bom = limpiar_bom(bom)
    print("RATE")
    rate = limpiar_rate(rate)

    # Comprobaciones sobre el resultado, antes de escribir nada.
    padres, comps = set(bom.ParentID), set(bom.ComponentID)
    productos = set(rate.ProductID)
    assert not bom.duplicated().any(), "quedan aristas padre/hijo repetidas"
    assert not rate.duplicated(["MachineId", "ProductID"]).any(), "quedan pares repetidos"
    assert not (bom.ParentID == bom.ComponentID).any(), "hay autorreferencias"
    assert padres == productos, (
        f"descuadre BOM/RATE: {len(padres - productos)} padres sin ruta, "
        f"{len(productos - padres)} productos sin BOM")

    with pd.ExcelWriter(FICHERO, engine="openpyxl") as xl:
        bom.to_excel(xl, sheet_name="BOM", index=False)
        rate.to_excel(xl, sheet_name="RATE", index=False)
    dar_formato(FICHERO)

    print(f"\n'{FICHERO}' reescrito.")
    print(f"  BOM : {len(bom)} aristas | {len(padres)} padres | {len(comps)} componentes")
    print(f"  RATE: {len(rate)} pares | {len(productos)} productos | "
          f"{rate.MachineId.nunique()} maquinas")
    print(f"  padres de BOM == productos de RATE: {padres == productos}")


if __name__ == "__main__":
    main()
